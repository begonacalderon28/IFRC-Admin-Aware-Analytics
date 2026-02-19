import re
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import fmean, pstdev

from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from api.analytics_access import get_analytics_access
from api.analytics_modules import (
    MODULE_AUDIENCE_INSIGHTS,
    MODULE_ENGAGEMENT_COMPARISON,
    MODULE_ENGAGEMENT_PERFORMANCE,
    MODULE_LIVE_SPIKES,
    MODULE_MAP_HEATMAP,
    MODULE_METADATA_LOOKUP,
    MODULE_OVERVIEW,
    MODULE_PLATFORM_ADOPTION,
    MODULE_TOP_COUNTRIES,
    MODULE_TOP_PAGES,
    MODULE_VIEWS_BY_DATE,
    get_available_modules,
    infer_role_profile,
)
from api.models import Country
from api.models import Event
from main.permissions import DenyGuestUserPermission

FACT_SHEET_NAME = "fact_views_daily_city"


def _find_dataset_path() -> Path:
    # Only use the new synthetic dataset.
    base_dir = Path(__file__).resolve().parents[1]
    candidates = [
        base_dir / "synthetic_ga_emergency_views.xlsx",
        base_dir.parent / "synthetic_ga_emergency_views.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("synthetic_ga_emergency_views.xlsx not found")


def _region_id_to_code(region_id: int) -> str | None:
    return {
        0: "africa",
        1: "americas",
        2: "asia-pacific",
        3: "europe",
        4: "middle-east-north-africa",
    }.get(region_id)


def _country_region_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.select_related("region").all():
        if country.name and country.region_id is not None:
            region_code = _region_id_to_code(country.region_id)
            if region_code:
                mapping[country.name.strip().lower()] = region_code
    return mapping


def _country_iso_region_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.select_related("region").all():
        if country.iso and country.region_id is not None:
            region_code = _region_id_to_code(country.region_id)
            if region_code:
                mapping[country.iso.strip().upper()] = region_code
    return mapping


def _country_iso_name_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.all():
        if country.iso and country.name:
            mapping[country.iso.strip().upper()] = country.name
    return mapping


def _country_iso_to_iso3_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.all():
        if country.iso and country.iso3:
            mapping[country.iso.strip().upper()] = country.iso3.strip().upper()
    return mapping


def _country_iso3_region_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.select_related("region").all():
        if country.iso3 and country.region_id is not None:
            region_code = _region_id_to_code(country.region_id)
            if region_code:
                mapping[country.iso3.strip().upper()] = region_code
    return mapping


def _country_iso3_name_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.all():
        if country.iso3 and country.name:
            mapping[country.iso3.strip().upper()] = country.name
    return mapping


def _country_name_title_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for country in Country.objects.all():
        if country.name:
            mapping[country.name.strip().lower()] = country.name
    return mapping


def _build_event_scope_map(event_ids: set[int]) -> dict[int, dict[str, set[str]]]:
    if not event_ids:
        return {}

    event_scope_map: dict[int, dict[str, set[str]]] = {}
    queryset = (
        Event.objects.filter(id__in=event_ids)
        .prefetch_related("regions", "countries__region")
    )
    for event in queryset:
        region_codes: set[str] = set()
        country_isos: set[str] = set()

        for region in event.regions.all():
            code = _region_id_to_code(region.name)
            if code:
                region_codes.add(code)

        for country in event.countries.all():
            if country.iso:
                country_isos.add(country.iso.strip().upper())
            if country.region_id is not None:
                code = _region_id_to_code(country.region_id)
                if code:
                    region_codes.add(code)

        event_scope_map[event.id] = {
            "regions": region_codes,
            "countries": country_isos,
        }

    return event_scope_map


def _infer_regions_from_emergency_name(
    emergency_name: str,
    country_name_region_map: dict[str, str],
    iso3_region_map: dict[str, str],
) -> set[str]:
    if not emergency_name:
        return set()

    regions: set[str] = set()
    lower_name = emergency_name.lower()

    iso3_match = re.match(r"^\s*([A-Za-z]{3})\s*:", emergency_name)
    if iso3_match:
        region = iso3_region_map.get(iso3_match.group(1).upper())
        if region:
            regions.add(region)

    region_markers = {
        "africa": "africa",
        "americas": "americas",
        "asia-pacific": "asia-pacific",
        "asia pacific": "asia-pacific",
        "europe": "europe",
        "mena": "middle-east-north-africa",
        "middle east": "middle-east-north-africa",
        "north africa": "middle-east-north-africa",
    }
    for marker, region_code in region_markers.items():
        if marker in lower_name:
            regions.add(region_code)

    for country_name, region_code in country_name_region_map.items():
        if len(country_name) < 4:
            continue
        if re.search(rf"\b{re.escape(country_name)}\b", lower_name):
            regions.add(region_code)

    return regions


def _is_active_emergency(row: dict) -> bool:
    value = str(row.get("is_active") or "").strip().lower()
    return value in {"yes", "true", "1", "y"}


def _parse_engagement_rate(value: str | None) -> float:
    try:
        return float(value or 0)
    except (ValueError, TypeError):
        return 0.0


def _parse_int(value, fallback: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _row_views(row: dict) -> int:
    return max(_parse_int(row.get("views"), 1), 0)


def _parse_query_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return None


def _parse_query_month(value: str | None) -> date | None:
    if not value:
        return None
    raw_value = str(value).strip()
    for fmt in ("%Y-%m", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(raw_value, fmt).date()
            return date(parsed.year, parsed.month, 1)
        except (TypeError, ValueError):
            continue
    return None


def _month_end(month_start: date) -> date:
    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    return next_month - timedelta(days=1)


def _load_xlsx_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[FACT_SHEET_NAME] if FACT_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[-1]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    output: list[dict] = []
    for row in rows:
        values = dict(zip(headers, row))
        page_path = values.get("page_path")
        if not page_path:
            continue

        date_value = values.get("date")
        if isinstance(date_value, datetime):
            date_value = date_value.date().isoformat()
        elif isinstance(date_value, date):
            date_value = date_value.isoformat()
        else:
            date_value = str(date_value or "")

        output.append(
            {
                "date": date_value,
                "fullPageUrl": str(page_path),
                "emergency_name": str(values.get("emergency_name") or ""),
                "country": str(values.get("viewer_country") or ""),
                "viewer_city": str(values.get("viewer_city") or ""),
                "views": _parse_int(values.get("views"), 0),
                "downloads": _parse_int(values.get("downloads"), 0),
                "engagementRate": str(values.get("avg_engagement_time_sec") or "0"),
                "is_active": str(values.get("is_active") or ""),
                "sessionSource": str(values.get("session_source") or ""),
                "new_returning_user": str(
                    values.get("new_vs_returning_user")
                    or values.get("new_vs_returning")
                    or values.get("new_returning_user")
                    or values.get("user type")
                    or values.get("user_type")
                    or values.get("new_user_vs_returning_user")
                    or ""
                ),
                "device": str(values.get("device") or ""),
                "browser": str(values.get("browser") or ""),
                "operatingSystemWithVersion": str(values.get("os") or ""),
                "emergency_id": _parse_int(values.get("emergency_id"), 0),
            }
        )
    return output


def _build_views_by_date(rows: list[dict], daily: bool = False) -> list[dict]:
    if not rows:
        return []
    buckets = Counter()
    for row in rows:
        raw_date = str(row.get("date") or "").strip()
        if daily:
            bucket = raw_date[:10] if len(raw_date) >= 10 else "unknown"
        else:
            bucket = raw_date[:7] if len(raw_date) >= 7 else "unknown"
        buckets[bucket] += _row_views(row)
    return [{"label": bucket, "views": count} for bucket, count in sorted(buckets.items())]


def _build_engagement_performance(rows: list[dict]) -> list[dict]:
    event_rows: dict[str, list[dict]] = {}
    for row in rows:
        event_id = str(_parse_int(row.get("emergency_id"), 0))
        if event_id == "0":
            continue
        event_rows.setdefault(event_id, []).append(row)

    latest_row_date: date | None = None
    for row in rows:
        row_date = _parse_query_date(str(row.get("date") or ""))
        if not row_date:
            continue
        if latest_row_date is None or row_date > latest_row_date:
            latest_row_date = row_date
    cutoff_date = (latest_row_date - timedelta(days=30)) if latest_row_date else None

    result: list[dict] = []
    for event_id, values in sorted(
        event_rows.items(),
        key=lambda item: sum(_row_views(r) for r in item[1]),
        reverse=True,
    ):
        weighted_views = sum(_row_views(r) for r in values)
        engagement_total = sum(_parse_engagement_rate(r.get("engagementRate")) * _row_views(r) for r in values)
        avg_engagement = engagement_total / max(weighted_views, 1)
        downloads = sum(_parse_int(r.get("downloads"), 0) for r in values)
        views_last_month = 0
        if cutoff_date is not None:
            for value in values:
                row_date = _parse_query_date(str(value.get("date") or ""))
                if row_date and row_date >= cutoff_date:
                    views_last_month += _row_views(value)
        emergency_name = next((str(v.get("emergency_name") or "").strip() for v in values if v.get("emergency_name")), "")
        result.append(
            {
                "event_id": event_id,
                "emergency_name": emergency_name or f"Emergency {event_id}",
                "page_url": f"https://go.ifrc.org/emergencies/{event_id}/details",
                "total_page_views": weighted_views,
                "views_last_month": views_last_month,
                "documents_download": downloads,
                "avg_engagement_time_sec": round(avg_engagement, 2),
            }
        )
    return result


def _build_metadata_lookup(rows: list[dict]) -> list[dict]:
    event_rows: dict[str, list[dict]] = {}
    for row in rows:
        event_id = str(row.get("emergency_id") or "")
        if not event_id or event_id == "0":
            continue
        event_rows.setdefault(event_id, []).append(row)

    payload = []
    for event_id, values in sorted(
        event_rows.items(),
        key=lambda item: sum(_row_views(r) for r in item[1]),
        reverse=True,
    ):
        total_views = sum(_row_views(v) for v in values)
        downloads = sum(_parse_int(v.get("downloads"), 0) for v in values)
        weighted_engagement = sum(_parse_engagement_rate(v.get("engagementRate")) * _row_views(v) for v in values)
        avg_engagement = weighted_engagement / max(total_views, 1)
        is_active = any(_is_active_emergency(v) for v in values)
        analytics_by_date = Counter()
        for value in values:
            date_key = str(value.get("date") or "")[:10]
            if not date_key:
                continue
            analytics_by_date[date_key] += _row_views(value)
        analytics_date = ""
        analytics_views = 0
        analytics_downloads = 0
        analytics_engagement_total = 0.0
        analytics_engagement_weight = 0
        if analytics_by_date:
            analytics_date = max(analytics_by_date.keys())
            analytics_views = analytics_by_date[analytics_date]
            analytics_downloads = sum(
                _parse_int(v.get("downloads"), 0)
                for v in values
                if str(v.get("date") or "")[:10] == analytics_date
            )
            for value in values:
                if str(value.get("date") or "")[:10] != analytics_date:
                    continue
                row_views = _row_views(value)
                analytics_engagement_total += _parse_engagement_rate(value.get("engagementRate")) * row_views
                analytics_engagement_weight += row_views
        analytics_spend_time = (
            analytics_engagement_total / max(analytics_engagement_weight, 1)
            if analytics_date
            else avg_engagement
        )
        source_counter = Counter()
        for value in values:
            source = value.get("sessionSource")
            if source:
                source_counter[source] += _row_views(value)
        top_source, top_source_views = ("", 0)
        if source_counter:
            top_source, top_source_views = source_counter.most_common(1)[0]
        top_source_pct = (top_source_views / max(total_views, 1)) * 100

        emergency_name = next((str(v.get("emergency_name") or "").strip() for v in values if v.get("emergency_name")), "")
        payload.append(
            {
                "event_id": event_id,
                "emergency_name": emergency_name or f"Emergency {event_id}",
                "page_url": f"https://go.ifrc.org/emergencies/{event_id}/details",
                "analytics_date": analytics_date,
                "views": analytics_views,
                "downloads": analytics_downloads if analytics_date else downloads,
                "spend_time_sec": round(analytics_spend_time, 2),
                "active_emergency": is_active,
                "primary_session_source": top_source,
                "primary_session_source_pct": round(top_source_pct, 1),
            }
        )
    return payload


def _build_live_spikes(
    rows: list[dict],
    rolling_window: int = 14,
    min_history_points: int = 5,
    min_non_zero_history_points: int = 3,
    sigma_multiplier: float = 3.5,
    min_absolute_views: int = 200,
    min_absolute_delta: int = 20,
    stddev_floor: float = 1.0,
) -> list[dict]:
    event_daily_views: dict[str, Counter] = {}
    event_names: dict[str, str] = {}

    for row in rows:
        event_id = str(_parse_int(row.get("emergency_id"), 0))
        if event_id == "0":
            continue
        row_date = _parse_query_date(str(row.get("date") or ""))
        if not row_date:
            continue
        if event_id not in event_daily_views:
            event_daily_views[event_id] = Counter()
        event_daily_views[event_id][row_date.isoformat()] += _row_views(row)
        if event_id not in event_names:
            emergency_name = str(row.get("emergency_name") or "").strip()
            if emergency_name:
                event_names[event_id] = emergency_name

    spikes: list[dict] = []
    for event_id, daily_counter in event_daily_views.items():
        ordered_points = sorted(daily_counter.items(), key=lambda item: item[0])
        if not ordered_points:
            continue

        min_day = _parse_query_date(ordered_points[0][0])
        max_day = _parse_query_date(ordered_points[-1][0])
        if not min_day or not max_day:
            continue

        # Build a dense daily series so gaps are treated as zero traffic.
        labels: list[str] = []
        values: list[int] = []
        current_day = min_day
        while current_day <= max_day:
            day_key = current_day.isoformat()
            labels.append(day_key)
            values.append(int(daily_counter.get(day_key, 0)))
            current_day += timedelta(days=1)

        for idx in range(len(values)):
            history_start = max(0, idx - rolling_window)
            history = values[history_start:idx]
            if len(history) < min_history_points:
                continue

            # Primary baseline: ignore zeros so long inactive periods do not collapse
            # mean/stddev and produce unstable infinite-style z-scores.
            non_zero_history = [value for value in history if value > 0]
            if len(non_zero_history) >= min_non_zero_history_points:
                baseline = non_zero_history
                baseline_mode = "non_zero"
            else:
                # Fallback when too few non-zero points exist.
                baseline = history
                baseline_mode = "all_values"

            history_mean = float(fmean(baseline))
            raw_stddev = float(pstdev(baseline)) if len(baseline) > 1 else 0.0
            effective_stddev = max(raw_stddev, stddev_floor)
            threshold = history_mean + (sigma_multiplier * effective_stddev)
            current_value = values[idx]
            z_score = (current_value - history_mean) / effective_stddev
            absolute_delta = current_value - history_mean
            if (
                z_score >= sigma_multiplier
                and current_value >= min_absolute_views
                and absolute_delta >= min_absolute_delta
            ):
                spikes.append(
                    {
                        "event_id": event_id,
                        "emergency_name": event_names.get(event_id, f"Emergency {event_id}"),
                        "date": labels[idx],
                        "views": current_value,
                        "baseline_mode": baseline_mode,
                        "baseline_mean": round(history_mean, 2),
                        "baseline_stddev": round(raw_stddev, 2),
                        "effective_stddev": round(effective_stddev, 2),
                        "threshold": round(threshold, 2),
                        "z_score": round(z_score, 2),
                    }
                )

    spikes.sort(key=lambda item: (item["z_score"], item["views"]), reverse=True)
    return spikes[:10]


def _build_platform_adoption(
    rows: list[dict],
    total_countries: int,
    event_scope_map: dict[int, dict[str, set[str]]],
    iso_name_map: dict[str, str],
    iso3_name_map: dict[str, str],
    country_name_region_map: dict[str, str],
) -> dict[str, object]:
    monthly_active_users = Counter()
    monthly_new_user_views = Counter()
    monthly_returning_user_views = Counter()
    monthly_countries = {}
    monthly_country_emergencies = {}
    event_ids = set()
    event_first_seen_month = {}
    event_country_name_cache: dict[int, set[str]] = {}

    for row in rows:
        raw_date = str(row.get("date") or "").strip()
        month_key = raw_date[:7] if len(raw_date) >= 7 else ""
        if not month_key:
            continue
        views = _row_views(row)

        new_returning = str(row.get("new_returning_user") or "").strip().lower()
        if "new" in new_returning:
            monthly_new_user_views[month_key] += views
            monthly_active_users[month_key] += views
        elif "return" in new_returning:
            monthly_returning_user_views[month_key] += views
            monthly_active_users[month_key] += views

        event_id = _parse_int(row.get("emergency_id"), 0)
        if event_id > 0:
            event_ids.add(event_id)
            if month_key:
                current_month = event_first_seen_month.get(event_id)
                if current_month is None or month_key < current_month:
                    event_first_seen_month[event_id] = month_key
            if month_key:
                if event_id not in event_country_name_cache:
                    owner_country_isos = event_scope_map.get(event_id, {}).get("countries", set())
                    owner_country_names = {
                        iso_name_map.get(iso, iso)
                        for iso in owner_country_isos
                        if iso
                    }
                    if not owner_country_names:
                        emergency_name = str(row.get("emergency_name") or "")
                        inferred_countries: set[str] = set()
                        iso3_match = re.match(r"^\s*([A-Za-z]{3})\s*:", emergency_name)
                        if iso3_match:
                            country_name = iso3_name_map.get(iso3_match.group(1).upper())
                            if country_name:
                                inferred_countries.add(country_name)
                        lowered_name = emergency_name.lower()
                        for country_name in country_name_region_map.keys():
                            if len(country_name) < 4:
                                continue
                            if re.search(rf"\b{re.escape(country_name)}\b", lowered_name):
                                inferred_countries.add(country_name.title())
                        owner_country_names = inferred_countries
                    event_country_name_cache[event_id] = owner_country_names
                owner_country_names = event_country_name_cache[event_id]
                if not owner_country_names:
                    continue
                for country_name in owner_country_names:
                    monthly_countries.setdefault(month_key, set()).add(country_name)
                    if month_key not in monthly_country_emergencies:
                        monthly_country_emergencies[month_key] = {}
                    if country_name not in monthly_country_emergencies[month_key]:
                        monthly_country_emergencies[month_key][country_name] = set()
                    monthly_country_emergencies[month_key][country_name].add(event_id)

    event_created_per_month = Counter()
    if event_ids:
        found_event_ids = set()
        for event in Event.objects.filter(id__in=event_ids).only("id", "created_at"):
            found_event_ids.add(event.id)
            if event.created_at:
                event_created_per_month[event.created_at.strftime("%Y-%m")] += 1
        # Fallback for synthetic/missing Event records: use first month seen in scoped analytics rows.
        missing_ids = event_ids - found_event_ids
        for event_id in missing_ids:
            fallback_month = event_first_seen_month.get(event_id)
            if fallback_month:
                event_created_per_month[fallback_month] += 1

    months = sorted(set(monthly_active_users.keys()) | set(event_created_per_month.keys()))
    monthly_breakdown = []
    for month in months:
        countries_count = len(monthly_countries.get(month, set()))
        countries_pct = round((countries_count / max(total_countries, 1)) * 100, 2)
        top_country = ""
        top_country_emergencies = 0
        month_country_events = monthly_country_emergencies.get(month, {})
        if month_country_events:
            top_country, event_set = max(
                month_country_events.items(),
                key=lambda item: len(item[1]),
            )
            top_country_emergencies = len(event_set)
        monthly_breakdown.append(
            {
                "month": month,
                "monthly_active_users": monthly_active_users.get(month, 0),
                "monthly_new_users": monthly_new_user_views.get(month, 0),
                "monthly_returning_users": monthly_returning_user_views.get(month, 0),
                "countries_publishing_pct": countries_pct,
                "emergencies_created": event_created_per_month.get(month, 0),
                "most_publishing_country": top_country,
                "most_publishing_country_emergencies": top_country_emergencies,
            }
        )

    return {
        "monthly_breakdown": monthly_breakdown,
    }


def _infer_countries_from_emergency_name(
    emergency_name: str,
    country_name_title_map: dict[str, str],
    iso3_name_map: dict[str, str],
) -> set[str]:
    if not emergency_name:
        return set()

    countries: set[str] = set()
    lower_name = emergency_name.lower()
    iso3_match = re.match(r"^\s*([A-Za-z]{3})\s*:", emergency_name)
    if iso3_match:
        country_name = iso3_name_map.get(iso3_match.group(1).upper())
        if country_name:
            countries.add(country_name)

    for country_name_lower, country_name in country_name_title_map.items():
        if len(country_name_lower) < 4:
            continue
        if re.search(rf"\b{re.escape(country_name_lower)}\b", lower_name):
            countries.add(country_name)
    return countries


def _build_engagement_comparison(
    rows: list[dict],
    role: str,
    role_regions: list[str],
    country_region_map: dict[str, str],
    country_name_title_map: dict[str, str],
    iso3_region_map: dict[str, str],
    event_scope_map: dict[int, dict[str, set[str]]],
    iso_name_map: dict[str, str],
    iso3_name_map: dict[str, str],
    mode_query: str | None,
    left_query: str | None,
    right_query: str | None,
    a_start_query: str | None,
    a_end_query: str | None,
    b_start_query: str | None,
    b_end_query: str | None,
) -> dict[str, object]:
    allowed_modes = ["country"] if role == "regional_im" else ["country", "region"]
    mode = mode_query if mode_query in allowed_modes else allowed_modes[0]
    allowed_region_set = set(role_regions)
    owner_scope_by_event_id: dict[int, dict[str, set[str]]] = {}

    def _owner_scope_for_row(row: dict) -> dict[str, set[str]]:
        event_id = _parse_int(row.get("emergency_id"), 0)
        if event_id in owner_scope_by_event_id:
            return owner_scope_by_event_id[event_id]

        owner_regions: set[str] = set()
        owner_countries: set[str] = set()
        event_scope = event_scope_map.get(event_id, {"regions": set(), "countries": set()})
        owner_regions.update(event_scope.get("regions", set()))
        owner_countries.update(
            iso_name_map.get(country_iso, country_iso)
            for country_iso in event_scope.get("countries", set())
            if iso_name_map.get(country_iso, country_iso)
        )

        emergency_name = str(row.get("emergency_name") or "")
        if not owner_countries:
            owner_countries.update(
                _infer_countries_from_emergency_name(
                    emergency_name,
                    country_name_title_map,
                    iso3_name_map,
                )
            )
        if not owner_regions:
            owner_regions.update(
                _infer_regions_from_emergency_name(
                    emergency_name,
                    country_region_map,
                    iso3_region_map,
                )
            )
            owner_regions.update(
                country_region_map.get(country_name.lower(), "")
                for country_name in owner_countries
                if country_region_map.get(country_name.lower(), "")
            )

        owner_scope_by_event_id[event_id] = {
            "regions": {region for region in owner_regions if region},
            "countries": {country for country in owner_countries if country},
        }
        return owner_scope_by_event_id[event_id]

    def _row_entities(row: dict, selected_mode: str) -> set[str]:
        scope = _owner_scope_for_row(row)
        if selected_mode == "region":
            regions = set(scope["regions"])
            if role == "regional_im":
                regions = regions.intersection(allowed_region_set)
            return regions
        countries = set(scope["countries"])
        if role == "regional_im":
            countries = {
                country_name
                for country_name in countries
                if country_region_map.get(country_name.lower()) in allowed_region_set
            }
        return countries

    option_values: set[str] = set()
    for row in rows:
        option_values.update(_row_entities(row, mode))
    options = sorted(option_values)

    selected_left = left_query if left_query in options else (options[0] if options else "")
    selected_right_default = options[1] if len(options) > 1 else selected_left
    selected_right = right_query if right_query in options else selected_right_default

    available_months = sorted(
        {
            str(row.get("date") or "")[:7]
            for row in rows
            if len(str(row.get("date") or "")) >= 7
        }
    )
    default_a_start = available_months[-1] if available_months else ""
    default_a_end = default_a_start
    default_b_start = available_months[-2] if len(available_months) >= 2 else default_a_start
    default_b_end = default_b_start

    period_a_start_month = a_start_query if a_start_query in available_months else default_a_start
    period_a_end_month = a_end_query if a_end_query in available_months else default_a_end
    period_b_start_month = b_start_query if b_start_query in available_months else default_b_start
    period_b_end_month = b_end_query if b_end_query in available_months else default_b_end

    period_a_start = _parse_query_month(period_a_start_month)
    period_a_end = _parse_query_month(period_a_end_month)
    period_b_start = _parse_query_month(period_b_start_month)
    period_b_end = _parse_query_month(period_b_end_month)
    if period_a_end:
        period_a_end = _month_end(period_a_end)
    if period_b_end:
        period_b_end = _month_end(period_b_end)

    if period_a_start and period_a_end and period_a_start > period_a_end:
        period_a_start, period_a_end = period_a_end, period_a_start
        period_a_start_month, period_a_end_month = period_a_end_month, period_a_start_month
    if period_b_start and period_b_end and period_b_start > period_b_end:
        period_b_start, period_b_end = period_b_end, period_b_start
        period_b_start_month, period_b_end_month = period_b_end_month, period_b_start_month

    def _build_metrics(entity: str, start: date | None, end: date | None) -> dict[str, float]:
        selected_rows = []
        for row in rows:
            if entity not in _row_entities(row, mode):
                continue
            row_date = _parse_query_date(str(row.get("date") or ""))
            if start and (not row_date or row_date < start):
                continue
            if end and (not row_date or row_date > end):
                continue
            selected_rows.append(row)
        total_views = sum(_row_views(row) for row in selected_rows)
        engagement_total = sum(_parse_engagement_rate(row.get("engagementRate")) * _row_views(row) for row in selected_rows)
        avg_engagement = engagement_total / max(total_views, 1)
        return {
            "total_page_views": total_views,
            "avg_engagement_time_sec": round(avg_engagement, 2),
        }

    return {
        "allowed_modes": allowed_modes,
        "mode": mode,
        "options": options,
        "available_months": available_months,
        "selected_left": selected_left,
        "selected_right": selected_right,
        "period_a_start": period_a_start_month,
        "period_a_end": period_a_end_month,
        "period_b_start": period_b_start_month,
        "period_b_end": period_b_end_month,
        "results": {
            "period_a": {
                "left": _build_metrics(selected_left, period_a_start, period_a_end) if selected_left else {},
                "right": _build_metrics(selected_right, period_a_start, period_a_end) if selected_right else {},
            },
            "period_b": {
                "left": _build_metrics(selected_left, period_b_start, period_b_end) if selected_left else {},
                "right": _build_metrics(selected_right, period_b_start, period_b_end) if selected_right else {},
            },
        },
    }


class AnalyticsView(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated, DenyGuestUserPermission)

    def get(self, request, *args, **kwargs):
        access = get_analytics_access(request.user)
        role_profile = infer_role_profile(access)

        scope = {
            "global": access["global_access"],
            "live": access["live_access"],
            "regions": access["region_codes"],
        }
        enforced_scope = {
            "global": scope["global"],
            "live": scope["live"],
            "regions": scope["regions"],
        }

        # IM officers (ops_im) active-emergency scoped.
        if role_profile["role"] == "ops_im":
            enforced_scope["global"] = False
            enforced_scope["live"] = True
            enforced_scope["regions"] = []

        dataset_path = _find_dataset_path()
        country_region = _country_region_map()
        iso3_region_map = _country_iso3_region_map()
        iso_region_map = _country_iso_region_map()
        iso_name_map = _country_iso_name_map()
        iso3_name_map = _country_iso3_name_map()
        country_name_title_map = _country_name_title_map()
        iso_to_iso3_map = _country_iso_to_iso3_map()
        rows = _load_xlsx_rows(dataset_path)
        event_ids = {_parse_int(r.get("emergency_id"), 0) for r in rows if _parse_int(r.get("emergency_id"), 0) > 0}
        event_scope_map = _build_event_scope_map(event_ids)
        fallback_scope_by_event_id: dict[int, dict[str, set[str]]] = {}
        start_date = _parse_query_date(request.query_params.get("start_date"))
        end_date = _parse_query_date(request.query_params.get("end_date"))
        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date

        filtered_rows: list[dict] = []
        for row in rows:
            country_iso = (row.get("country") or "").strip().upper()
            row["country_iso"] = country_iso
            row["country"] = iso_name_map.get(country_iso, country_iso)
            event_id = _parse_int(row.get("emergency_id"), 0)
            event_scope = event_scope_map.get(event_id, {"regions": set(), "countries": set()})
            if event_id > 0 and not event_scope["regions"]:
                if event_id not in fallback_scope_by_event_id:
                    inferred_regions = _infer_regions_from_emergency_name(
                        str(row.get("emergency_name") or ""),
                        country_region,
                        iso3_region_map,
                    )
                    fallback_scope_by_event_id[event_id] = {"regions": inferred_regions, "countries": set()}
                event_scope = fallback_scope_by_event_id[event_id]

            if not enforced_scope["global"]:
                if enforced_scope["regions"]:
                    if not event_scope["regions"].intersection(enforced_scope["regions"]):
                        continue
                elif enforced_scope["live"]:
                    if not _is_active_emergency(row):
                        continue
                else:
                    continue

            if enforced_scope["live"] and not enforced_scope["global"] and not _is_active_emergency(row):
                continue

            filtered_rows.append(row)

        views_by_date_rows: list[dict] = []
        for row in filtered_rows:
            row_date = _parse_query_date(str(row.get("date") or ""))
            if start_date and (not row_date or row_date < start_date):
                continue
            if end_date and (not row_date or row_date > end_date):
                continue
            views_by_date_rows.append(row)

        available_modules = get_available_modules(role_profile["role"])
        total_visits = sum(_row_views(r) for r in filtered_rows)
        top_pages_counter = Counter()
        top_countries_counter = Counter()
        for row in filtered_rows:
            views = _row_views(row)
            page = row.get("fullPageUrl")
            country = row.get("country")
            if page:
                top_pages_counter[page] += views
            if country:
                top_countries_counter[country] += views
        top_pages = top_pages_counter.most_common(10)
        top_countries = top_countries_counter.most_common(10)
        emergency_rows = [r for r in filtered_rows if _is_active_emergency(r)]
        module_data: dict[str, object] = {}

        if MODULE_OVERVIEW in available_modules:
            module_data[MODULE_OVERVIEW] = {
                "total_visits": total_visits,
                "total_emergency_views": len(emergency_rows),
                "unique_countries": len({r.get("country") for r in filtered_rows if r.get("country")}),
            }
        if MODULE_VIEWS_BY_DATE in available_modules:
            all_months_series = _build_views_by_date(filtered_rows)
            use_daily_buckets = bool(
                start_date
                and end_date
                and start_date.year == end_date.year
                and start_date.month == end_date.month
            )
            module_data[MODULE_VIEWS_BY_DATE] = {
                "series": _build_views_by_date(views_by_date_rows, daily=use_daily_buckets),
                "available_labels": [item["label"] for item in all_months_series],
            }
        if MODULE_TOP_PAGES in available_modules:
            module_data[MODULE_TOP_PAGES] = top_pages
        if MODULE_TOP_COUNTRIES in available_modules:
            module_data[MODULE_TOP_COUNTRIES] = top_countries
        if MODULE_MAP_HEATMAP in available_modules:
            country_iso3_counter = Counter()
            country_views_counter = Counter()
            city_views_by_country: dict[str, Counter] = {}
            can_city_drilldown = role_profile["role"] == "regional_im"
            for row in filtered_rows:
                country_name = str(row.get("country") or "").strip()
                if country_name:
                    row_views = _row_views(row)
                    country_views_counter[country_name] += row_views
                    city_name = str(row.get("viewer_city") or "").strip()
                    if can_city_drilldown and city_name:
                        if country_name not in city_views_by_country:
                            city_views_by_country[country_name] = Counter()
                        city_views_by_country[country_name][city_name] += row_views
                row_iso = (row.get("country_iso") or "").strip().upper()
                iso3 = iso_to_iso3_map.get(row_iso)
                if iso3:
                    country_iso3_counter[iso3] += _row_views(row)
            module_data[MODULE_MAP_HEATMAP] = {
                "country_views": top_countries,
                "country_views_all": country_views_counter.most_common(),
                "country_views_iso3": [
                    {"iso3": iso3, "views": views}
                    for iso3, views in country_iso3_counter.items()
                ],
                "can_city_drilldown": can_city_drilldown,
                "city_views_by_country": {
                    country: city_counter.most_common(30)
                    for country, city_counter in city_views_by_country.items()
                },
            }
        if MODULE_ENGAGEMENT_PERFORMANCE in available_modules:
            module_data[MODULE_ENGAGEMENT_PERFORMANCE] = _build_engagement_performance(filtered_rows)
        if MODULE_AUDIENCE_INSIGHTS in available_modules:
            module_data[MODULE_AUDIENCE_INSIGHTS] = {
                "by_source": Counter(r.get("sessionSource") for r in filtered_rows if r.get("sessionSource")).most_common(8),
                "by_device": Counter(r.get("device") for r in filtered_rows if r.get("device")).most_common(8),
                "by_browser": Counter(r.get("browser") for r in filtered_rows if r.get("browser")).most_common(8),
                "by_os": Counter(
                    r.get("operatingSystemWithVersion")
                    for r in filtered_rows
                    if r.get("operatingSystemWithVersion")
                ).most_common(8),
            }
        if MODULE_LIVE_SPIKES in available_modules:
            module_data[MODULE_LIVE_SPIKES] = _build_live_spikes(filtered_rows)
        if MODULE_PLATFORM_ADOPTION in available_modules:
            module_data[MODULE_PLATFORM_ADOPTION] = _build_platform_adoption(
                filtered_rows,
                total_countries=len(iso_region_map),
                event_scope_map=event_scope_map,
                iso_name_map=iso_name_map,
                iso3_name_map=iso3_name_map,
                country_name_region_map=country_region,
            )
        if MODULE_ENGAGEMENT_COMPARISON in available_modules:
            module_data[MODULE_ENGAGEMENT_COMPARISON] = _build_engagement_comparison(
                filtered_rows,
                role=role_profile["role"],
                role_regions=enforced_scope["regions"],
                country_region_map=country_region,
                country_name_title_map=country_name_title_map,
                iso3_region_map=iso3_region_map,
                event_scope_map=event_scope_map,
                iso_name_map=iso_name_map,
                iso3_name_map=iso3_name_map,
                mode_query=request.query_params.get("cmp_mode"),
                left_query=request.query_params.get("cmp_left"),
                right_query=request.query_params.get("cmp_right"),
                a_start_query=request.query_params.get("cmp_a_start"),
                a_end_query=request.query_params.get("cmp_a_end"),
                b_start_query=request.query_params.get("cmp_b_start"),
                b_end_query=request.query_params.get("cmp_b_end"),
            )
        if MODULE_METADATA_LOOKUP in available_modules:
            module_data[MODULE_METADATA_LOOKUP] = _build_metadata_lookup(filtered_rows)

        return Response({
            "contract_version": 1,
            "role_profile": {
                **role_profile,
                "content_scope": {
                    "global": enforced_scope["global"],
                    "regions": enforced_scope["regions"],
                    "live": enforced_scope["live"],
                },
                "audience_scope": {"global": True},
            },
            "scope": enforced_scope,
            "filters_applied": {
                "start_date": start_date.isoformat() if start_date else None,
                "end_date": end_date.isoformat() if end_date else None,
            },
            "available_modules": available_modules,
            "module_data": module_data,
            "summary": {
                "total_visits": total_visits,
                "top_pages": top_pages,
                "top_countries": top_countries,
            },
        })
